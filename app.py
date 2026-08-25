from __future__ import annotations

import hashlib
import json
import os
import re
import sqlite3
import zipfile
from contextlib import closing
from datetime import date, datetime
from io import StringIO
from pathlib import Path
from typing import Iterable
from uuid import uuid4

from flask import (
	Flask,
	flash,
	g,
	redirect,
	render_template,
	send_file,
	request,
	url_for,
)
from ofxparse import OfxParser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from werkzeug.utils import secure_filename

try:
	import psycopg
	from psycopg.rows import dict_row
except ImportError:  # pragma: no cover - optional for local SQLite-only runs
	psycopg = None
	dict_row = None

if psycopg is None:
	DB_INTEGRITY_ERROR = (sqlite3.IntegrityError,)
else:
	DB_INTEGRITY_ERROR = (sqlite3.IntegrityError, psycopg.IntegrityError)


BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
UPLOAD_DIR = INSTANCE_DIR / "uploads"
SPEEDPOINT_RUNS_DIR = UPLOAD_DIR / "speedpoint_runs"
DB_PATH = INSTANCE_DIR / "accounting.db"
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)
ACCOUNT_NAMES_PATH = BASE_DIR / "Account Names.xlsx"

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
SPEEDPOINT_UPLOAD_EXTENSIONS = {".xlsx"}
PAYMENT_COLUMNS = (2, 4)
PAYMENT_RANGE = range(18, 34)
SPEEDPOINT_HIGHLIGHT = PatternFill(fill_type="solid", fgColor="FF86B9")
BANK_HIGHLIGHT = PatternFill(fill_type="solid", fgColor="FFFF00")

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-accounting-secret")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

app_name = os.environ.get('APP_NAME', 'Company Name')

@app.context_processor
def inject_global_vars():
    return dict(app_name=app_name)


class QueryResult:
	def __init__(self, cursor, lastrowid: int | None = None):
		self._cursor = cursor
		self.lastrowid = lastrowid
		self.rowcount = getattr(cursor, "rowcount", -1)

	def fetchone(self):
		return self._cursor.fetchone()

	def fetchall(self):
		return self._cursor.fetchall()

	def __getattr__(self, name: str):
		return getattr(self._cursor, name)


def _translate_sql(sql: str) -> str:
	if not USE_POSTGRES:
		return sql
	return sql.replace("?", "%s")


def _postgre_insert_with_returning(sql: str) -> bool:
	return USE_POSTGRES and sql.lstrip().lower().startswith("insert") and "returning" not in sql.lower()


def normalize_text(value: object) -> str:
	text = str(value or "").strip().upper()
	return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def normalize_tokens(value: object) -> list[str]:
	text = normalize_text(value)
	if not text:
		return []
	return text.split()


def matchable_tokens(value: object) -> list[str]:
	return [token for token in normalize_tokens(value) if not token.isdigit()]


def significant_tokens(value: object) -> list[str]:
	stop_words = {
		"AND",
		"FOR",
		"FROM",
		"IN",
		"OF",
		"ON",
		"OR",
		"THE",
		"TO",
		"WITH",
		"ACCOUNT",
		"ACCOUNTS",
		"CASH",
		"PAYMENT",
		"PAYMENTS",
		"SERVICE",
		"SERVICES",
		"TRANSFER",
		"TRANSACTIONS",
		"WAGES",
	}
	return [
		token
		for token in normalize_tokens(value)
		if len(token) >= 4 and token not in stop_words
	]


def account_match_score(account_name: object, description_tokens: set[str]) -> int:
	account_all_tokens = matchable_tokens(account_name)
	if not account_all_tokens or not description_tokens:
		return 0

	matched_all_tokens = [token for token in account_all_tokens if token in description_tokens]
	if len(matched_all_tokens) >= 2:
		return len(matched_all_tokens) * 4 + len(account_all_tokens)

	account_tokens = significant_tokens(account_name)
	if not account_tokens:
		return 0

	matched_tokens = [token for token in account_tokens if token in description_tokens]
	if not matched_tokens:
		return 0

	if len(account_tokens) == 1:
		token = matched_tokens[0]
		return 5 if len(token) >= 4 else 0

	token = matched_tokens[0]
	if len(token) >= 7:
		return 4
	return 0


def parse_money(value: object) -> float | None:
	if value in (None, ""):
		return None
	if isinstance(value, str):
		cleaned = value.replace(" ", "").replace(",", "")
		if not cleaned:
			return None
		try:
			number = float(cleaned)
		except ValueError:
			return None
	else:
		try:
			number = float(value)
		except (TypeError, ValueError):
			return None
	if abs(number) < 1e-9:
		return None
	return round(number, 2)


def parse_filename_date(filename: str) -> str:
	name = Path(filename).stem
	patterns = [
		r"(?P<day>\d{1,2})[\s_\-\.]+(?P<month>[A-Za-z]+)[\s_\-\.]+(?P<year>\d{4})",
		r"(?P<day>\d{1,2})[\s_\-\.]+(?P<month>\d{1,2})[\s_\-\.]+(?P<year>\d{4})",
	]
	month_lookup = {
		"JAN": 1,
		"JANUARY": 1,
		"FEB": 2,
		"FEBRUARY": 2,
		"MAR": 3,
		"MARCH": 3,
		"APR": 4,
		"APRIL": 4,
		"MAY": 5,
		"JUN": 6,
		"JUNE": 6,
		"JUL": 7,
		"JULY": 7,
		"AUG": 8,
		"AUGUST": 8,
		"SEP": 9,
		"SEPT": 9,
		"SEPTEMBER": 9,
		"OCT": 10,
		"OCTOBER": 10,
		"NOV": 11,
		"NOVEMBER": 11,
		"DEC": 12,
		"DECEMBER": 12,
	}
	for pattern in patterns:
		match = re.search(pattern, name, flags=re.IGNORECASE)
		if not match:
			continue
		day = int(match.group("day"))
		year = int(match.group("year"))
		month_text = match.group("month")
		if month_text.isdigit():
			month = int(month_text)
		else:
			month = month_lookup.get(month_text.strip().upper())
		if month is None:
			continue
		return datetime(year, month, day).date().isoformat()
	raise ValueError(
		f"Could not find a date in the filename '{filename}'. "
		"Expected something like 'Control Sheet Daily 01 JULY 2026.xlsx'."
	)


def allowed_file(filename: str) -> bool:
	return Path(filename).suffix.lower() in SUPPORTED_EXTENSIONS


def get_connection() -> sqlite3.Connection:
	if "db" not in g:
		if USE_POSTGRES:
			if psycopg is None or dict_row is None:
				raise RuntimeError("DATABASE_URL is set, but psycopg is not installed.")
			conn = psycopg.connect(DATABASE_URL, row_factory=dict_row)
		else:
			conn = sqlite3.connect(DB_PATH)
			conn.row_factory = sqlite3.Row
			conn.execute("PRAGMA foreign_keys = ON")
		g.db = conn
	return g.db


def query_one(sql: str, params: tuple = ()) -> sqlite3.Row | None:
	return get_connection().execute(_translate_sql(sql), params).fetchone()


def query_all(sql: str, params: tuple = ()) -> list[sqlite3.Row]:
	return get_connection().execute(_translate_sql(sql), params).fetchall()


def execute(sql: str, params: tuple = ()) -> sqlite3.Cursor:
	conn = get_connection()
	translated_sql = _translate_sql(sql)
	lastrowid = None
	if _postgre_insert_with_returning(sql):
		translated_sql = f"{translated_sql.rstrip().rstrip(';')} RETURNING id"
	cur = conn.execute(translated_sql, params)
	if _postgre_insert_with_returning(sql):
		row = cur.fetchone()
		lastrowid = row["id"] if row else None
	elif not USE_POSTGRES:
		lastrowid = None if getattr(cur, "rowcount", -1) == 0 else cur.lastrowid
	conn.commit()
	return QueryResult(cur, lastrowid)


def initialize_storage() -> None:
	INSTANCE_DIR.mkdir(exist_ok=True)
	UPLOAD_DIR.mkdir(exist_ok=True)
	SPEEDPOINT_RUNS_DIR.mkdir(exist_ok=True)
	if USE_POSTGRES:
		with closing(psycopg.connect(DATABASE_URL, row_factory=dict_row)) as conn:
			for statement in postgres_schema_statements():
				conn.execute(statement)
			columns = postgres_table_columns(conn, "import_batches")
			if "source_type" not in columns:
				conn.execute("ALTER TABLE import_batches ADD COLUMN source_type TEXT NOT NULL DEFAULT 'cash'")
			conn.execute(
				"UPDATE import_batches SET source_type = 'cash' WHERE source_type IS NULL OR TRIM(source_type) = ''"
			)
			conn.commit()
		return

	with closing(sqlite3.connect(DB_PATH)) as conn:
		conn.row_factory = sqlite3.Row
		conn.execute("PRAGMA foreign_keys = ON")
		conn.executescript("\n".join(sqlite_schema_statements()))
		columns = {row["name"] for row in conn.execute("PRAGMA table_info(import_batches)")}
		if "source_type" not in columns:
			conn.execute("ALTER TABLE import_batches ADD COLUMN source_type TEXT NOT NULL DEFAULT 'cash'")
		conn.execute(
			"UPDATE import_batches SET source_type = 'cash' WHERE source_type IS NULL OR TRIM(source_type) = ''"
		)
		conn.commit()


def postgres_table_columns(conn, table_name: str) -> set[str]:
	rows = conn.execute(
		"""
		SELECT column_name
		FROM information_schema.columns
		WHERE table_schema = current_schema()
		  AND table_name = %s
		""",
		(table_name,),
	).fetchall()
	return {row["column_name"] for row in rows}


def sqlite_schema_statements() -> list[str]:
	return [
		"""
		CREATE TABLE IF NOT EXISTS accounts (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			name TEXT NOT NULL UNIQUE,
			source_sheet TEXT,
			created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
		);
		""",
		"""
		CREATE TABLE IF NOT EXISTS supplier_rules (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			supplier_key TEXT NOT NULL UNIQUE,
			supplier_name TEXT NOT NULL,
			account_id INTEGER NOT NULL,
			created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
			FOREIGN KEY (account_id) REFERENCES accounts (id) ON DELETE CASCADE
		);
		""",
		"""
		CREATE TABLE IF NOT EXISTS import_batches (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
			source_folder TEXT,
			source_type TEXT NOT NULL DEFAULT 'cash',
			status TEXT NOT NULL DEFAULT 'open'
		);
		""",
		"""
		CREATE TABLE IF NOT EXISTS import_files (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			batch_id INTEGER NOT NULL,
			original_filename TEXT NOT NULL,
			stored_filename TEXT NOT NULL,
			file_hash TEXT NOT NULL UNIQUE,
			file_date TEXT NOT NULL,
			created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
			FOREIGN KEY (batch_id) REFERENCES import_batches (id) ON DELETE CASCADE
		);
		""",
		"""
		CREATE TABLE IF NOT EXISTS staged_payments (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			batch_id INTEGER NOT NULL,
			file_id INTEGER NOT NULL,
			payment_date TEXT NOT NULL,
			supplier_name TEXT NOT NULL,
			supplier_key TEXT NOT NULL,
			amount_paid REAL NOT NULL,
			source_sheet TEXT NOT NULL,
			source_row INTEGER NOT NULL,
			source_column INTEGER NOT NULL,
			account_id INTEGER,
			created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
			UNIQUE (file_id, source_sheet, source_row, source_column),
			FOREIGN KEY (batch_id) REFERENCES import_batches (id) ON DELETE CASCADE,
			FOREIGN KEY (file_id) REFERENCES import_files (id) ON DELETE CASCADE,
			FOREIGN KEY (account_id) REFERENCES accounts (id) ON DELETE SET NULL
		);
		""",
		"""
		CREATE TABLE IF NOT EXISTS payments (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			payment_date TEXT NOT NULL,
			supplier_name TEXT NOT NULL,
			amount_paid REAL NOT NULL,
			account_id INTEGER NOT NULL,
			source_file_name TEXT NOT NULL,
			source_file_hash TEXT NOT NULL,
			source_sheet TEXT NOT NULL,
			source_row INTEGER NOT NULL,
			source_column INTEGER NOT NULL,
			created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
			UNIQUE (source_file_hash, source_sheet, source_row, source_column),
			FOREIGN KEY (account_id) REFERENCES accounts (id) ON DELETE RESTRICT
		);
		""",
	]


def postgres_schema_statements() -> list[str]:
	return [
		"""
		CREATE TABLE IF NOT EXISTS accounts (
			id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
			name TEXT NOT NULL UNIQUE,
			source_sheet TEXT,
			created_at TEXT NOT NULL DEFAULT (CURRENT_TIMESTAMP::text)
		)
		""",
		"""
		CREATE TABLE IF NOT EXISTS supplier_rules (
			id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
			supplier_key TEXT NOT NULL UNIQUE,
			supplier_name TEXT NOT NULL,
			account_id INTEGER NOT NULL,
			created_at TEXT NOT NULL DEFAULT (CURRENT_TIMESTAMP::text),
			FOREIGN KEY (account_id) REFERENCES accounts (id) ON DELETE CASCADE
		)
		""",
		"""
		CREATE TABLE IF NOT EXISTS import_batches (
			id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
			created_at TEXT NOT NULL DEFAULT (CURRENT_TIMESTAMP::text),
			source_folder TEXT,
			source_type TEXT NOT NULL DEFAULT 'cash',
			status TEXT NOT NULL DEFAULT 'open'
		)
		""",
		"""
		CREATE TABLE IF NOT EXISTS import_files (
			id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
			batch_id INTEGER NOT NULL,
			original_filename TEXT NOT NULL,
			stored_filename TEXT NOT NULL,
			file_hash TEXT NOT NULL UNIQUE,
			file_date TEXT NOT NULL,
			created_at TEXT NOT NULL DEFAULT (CURRENT_TIMESTAMP::text),
			FOREIGN KEY (batch_id) REFERENCES import_batches (id) ON DELETE CASCADE
		)
		""",
		"""
		CREATE TABLE IF NOT EXISTS staged_payments (
			id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
			batch_id INTEGER NOT NULL,
			file_id INTEGER NOT NULL,
			payment_date TEXT NOT NULL,
			supplier_name TEXT NOT NULL,
			supplier_key TEXT NOT NULL,
			amount_paid REAL NOT NULL,
			source_sheet TEXT NOT NULL,
			source_row INTEGER NOT NULL,
			source_column INTEGER NOT NULL,
			account_id INTEGER,
			created_at TEXT NOT NULL DEFAULT (CURRENT_TIMESTAMP::text),
			UNIQUE (file_id, source_sheet, source_row, source_column),
			FOREIGN KEY (batch_id) REFERENCES import_batches (id) ON DELETE CASCADE,
			FOREIGN KEY (file_id) REFERENCES import_files (id) ON DELETE CASCADE,
			FOREIGN KEY (account_id) REFERENCES accounts (id) ON DELETE SET NULL
		)
		""",
		"""
		CREATE TABLE IF NOT EXISTS payments (
			id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
			payment_date TEXT NOT NULL,
			supplier_name TEXT NOT NULL,
			amount_paid REAL NOT NULL,
			account_id INTEGER NOT NULL,
			source_file_name TEXT NOT NULL,
			source_file_hash TEXT NOT NULL,
			source_sheet TEXT NOT NULL,
			source_row INTEGER NOT NULL,
			source_column INTEGER NOT NULL,
			created_at TEXT NOT NULL DEFAULT (CURRENT_TIMESTAMP::text),
			UNIQUE (source_file_hash, source_sheet, source_row, source_column),
			FOREIGN KEY (account_id) REFERENCES accounts (id) ON DELETE RESTRICT
		)
		""",
	]


def sync_account_names() -> dict[str, int]:
    created: dict[str, int] = {}
    if not ACCOUNT_NAMES_PATH.exists():
        return created

    wb = load_workbook(ACCOUNT_NAMES_PATH, data_only=True, read_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value in (None, "", "."):
                continue
            name = str(cell.value).strip()
            if not name:
                continue
            existing = find_account_by_normalized_name(name)
            if existing:
                created[name] = existing["id"]
                continue
            cursor = execute(
                "INSERT INTO accounts (name, source_sheet) VALUES (?, ?) ON CONFLICT (name) DO NOTHING",
                (name, ws.title),
            )
            account_id = cursor.lastrowid
            if not account_id:
                row = query_one("SELECT id FROM accounts WHERE name = ?", (name,))
                account_id = row["id"] if row else None
            if account_id:
                created[name] = account_id
    return created


def find_account_by_normalized_name(name: str, exclude_account_id: int | None = None) -> sqlite3.Row | None:
    target = normalize_text(name)
    if not target:
        return None
    for account in query_all("SELECT id, name FROM accounts"):
        if exclude_account_id is not None and account["id"] == exclude_account_id:
            continue
        if normalize_text(account["name"]) == target:
            return account
    return None


def canonical_account_name(name: str) -> str:
    return str(name or "").strip().upper()


def ensure_rule(supplier_name: str, account_id: int) -> None:
	key = normalize_text(supplier_name)
	execute(
		"""
		INSERT INTO supplier_rules (supplier_key, supplier_name, account_id)
		VALUES (?, ?, ?)
		ON CONFLICT(supplier_key) DO UPDATE SET
			supplier_name = excluded.supplier_name,
			account_id = excluded.account_id
		""",
		(key, supplier_name.strip(), account_id),
	)


def resolve_account_for_supplier(supplier_name: str) -> sqlite3.Row | None:
	supplier_key = normalize_text(supplier_name)
	rule = query_one(
		"""
		SELECT supplier_rules.account_id, accounts.name
		FROM supplier_rules
		JOIN accounts ON accounts.id = supplier_rules.account_id
		WHERE supplier_rules.supplier_key = ?
		""",
		(supplier_key,),
	)
	if rule:
		return rule

	description_tokens = set(matchable_tokens(supplier_name))
	best_account = None
	best_score = 0
	for account in query_all("SELECT id, name FROM accounts"):
		score = account_match_score(account["name"], description_tokens)
		if score > best_score:
			best_score = score
			best_account = account
	if best_account and best_score > 0:
		return best_account

	exact = find_account_by_normalized_name(supplier_name)
	if exact:
		ensure_rule(supplier_name, exact["id"])
		return exact
	return None


def sha256_file(path: Path) -> str:
	digest = hashlib.sha256()
	with path.open("rb") as handle:
		for chunk in iter(lambda: handle.read(8192), b""):
			digest.update(chunk)
	return digest.hexdigest()


def pick_payment_sheet(workbook):
	for worksheet in workbook.worksheets:
		if normalize_text(worksheet.title) == "CASH UP":
			return worksheet
	return workbook.worksheets[0]


def compact_text(value: object) -> str:
	return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_bank_transaction(value: object) -> str:
	return compact_text(value).upper()


def format_download_name(original_name: str, strip_leading_number: bool = False) -> str:
	path = Path(original_name)
	stem = path.stem.replace("_", " ")
	if strip_leading_number:
		stem = re.sub(r"^\s*\d+[\s._-]*", "", stem)
	stem = re.sub(r"\s+", " ", stem).strip()
	if not stem:
		stem = "download"
	suffix = path.suffix or ".xlsx"
	return f"{stem}{suffix}"


def parse_speedpoint_bank_transaction(transaction: object) -> dict[str, object] | None:
	text = normalize_bank_transaction(transaction)
	if not text:
		return None

	speedpoint_match = re.match(r"^SPEEDPOINT\s*(?P<terminal>\d+)\s+(?P<batch>\d+)$", text, flags=re.IGNORECASE)
	if speedpoint_match:
		terminal = int(speedpoint_match.group("terminal"))
		terminal_map = {
			946389: (2, 3),
			946390: (4, 5),
		}
		speedpoint_column, amount_column = terminal_map.get(terminal, (2, 3))
		return {
			"transaction_type": "SPEEDPOINT",
			"bank_terminal": speedpoint_match.group("terminal"),
			"batch_number": int(speedpoint_match.group("batch")),
			"speedpoint_terminal": terminal,
			"speedpoint_column": speedpoint_column,
			"amount_column": amount_column,
		}

	nedlnk_match = re.search(r"NEDLNK\s+DP(?:\s+(?P<terminal>\d+))?(?:\s+(?P<batch>\d+))?", text, flags=re.IGNORECASE)
	if nedlnk_match:
		return {
			"transaction_type": "NEDLNK DP",
			"bank_terminal": nedlnk_match.group("terminal"),
			"batch_number": int(nedlnk_match.group("batch")) if nedlnk_match.group("batch") else None,
		}

	return None


def pick_bank_statement_sheet(workbook):
	for worksheet in workbook.worksheets:
		for row_number in range(1, min(worksheet.max_row, 5) + 1):
			values = [normalize_text(worksheet.cell(row_number, column).value) for column in (1, 2, 4)]
			if values == ["DATE", "TRANSACTIONS", "CREDIT"]:
				return worksheet
	return workbook.worksheets[0]


def pick_speedpoint_sheet(workbook):
	for worksheet in workbook.worksheets:
		if normalize_text(worksheet.title) in {"SPEED POINTS", "SPEEDPOINT", "SPEEDPOINTS"}:
			return worksheet
	return workbook.worksheets[0]


def parse_amount(value: object) -> float | None:
	number = parse_money(value)
	if number is None:
		return None
	return round(number, 2)


def find_speedpoint_match(
	worksheet,
	terminal_column: int,
	batch_number: int,
	expected_amount: float,
) -> tuple[int | None, float | None, bool]:
	candidate_row: int | None = None
	candidate_amount: float | None = None
	for row_number in range(1, worksheet.max_row + 1):
		batch_value = worksheet.cell(row=row_number, column=terminal_column).value
		if batch_value in (None, ""):
			continue
		try:
			current_batch = int(float(str(batch_value).strip()))
		except (TypeError, ValueError):
			continue
		if current_batch != batch_number:
			continue

		amount_value = parse_amount(worksheet.cell(row=row_number, column=terminal_column + 1).value)
		if amount_value == expected_amount:
			return row_number, amount_value, True
		if candidate_row is None:
			candidate_row = row_number
			candidate_amount = amount_value
	return candidate_row, candidate_amount, False


def find_speedpoint_total_match(
	worksheet,
	expected_amount: float,
	start_column: int = 7,
	end_column: int = 15,
) -> tuple[int | None, float | None, bool]:
	candidate_row: int | None = None
	candidate_total: float | None = None
	for row_number in range(1, worksheet.max_row + 1):
		row_total = 0.0
		has_amount = False
		for column_number in range(start_column, end_column + 1):
			amount = parse_amount(worksheet.cell(row=row_number, column=column_number).value)
			if amount is None:
				continue
			row_total += amount
			has_amount = True
		if not has_amount:
			continue
		row_total = round(row_total, 2)
		if row_total == expected_amount:
			return row_number, row_total, True
		if candidate_row is None:
			candidate_row = row_number
			candidate_total = row_total
	return candidate_row, candidate_total, False


def highlight_speedpoint_match(worksheet, row_number: int, terminal_column: int) -> None:
	worksheet.cell(row=row_number, column=terminal_column + 1).fill = SPEEDPOINT_HIGHLIGHT


def highlight_bank_match(worksheet, row_number: int, credit_column: int) -> None:
	worksheet.cell(row=row_number, column=credit_column).fill = BANK_HIGHLIGHT


def highlight_speedpoint_range(worksheet, row_number: int, start_column: int, end_column: int) -> None:
	for column_number in range(start_column, end_column + 1):
		worksheet.cell(row=row_number, column=column_number).fill = SPEEDPOINT_HIGHLIGHT


def build_speedpoint_result_dir(run_id: str) -> Path:
	run_dir = SPEEDPOINT_RUNS_DIR / run_id
	run_dir.mkdir(parents=True, exist_ok=True)
	return run_dir


def build_result_zip(run_dir: Path, manifest: dict[str, object]) -> Path:
	zip_path = run_dir / manifest["zip_filename"]
	with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
		for item in manifest["edited_files"]:
			file_path = run_dir / item["stored_name"]
			archive.write(file_path, arcname=item["download_name"])
	return zip_path


def process_speedpoint_reconciliation(speedpoint_storage, bank_storages: list) -> dict[str, object]:
	run_id = uuid4().hex
	run_dir = build_speedpoint_result_dir(run_id)

	speedpoint_original_name = Path(speedpoint_storage.filename).name
	speedpoint_safe_name = secure_filename(speedpoint_original_name) or "speedpoint.xlsx"
	speedpoint_input_path = run_dir / f"input_{speedpoint_safe_name}"
	speedpoint_storage.save(speedpoint_input_path)

	speedpoint_workbook = load_workbook(speedpoint_input_path)
	speedpoint_sheet = pick_speedpoint_sheet(speedpoint_workbook)

	unresolved_rows: list[dict[str, object]] = []
	total_matches = 0
	edited_files: list[dict[str, str]] = []

	for index, bank_storage in enumerate(bank_storages, start=1):
		bank_original_name = Path(bank_storage.filename).name
		bank_safe_name = secure_filename(bank_original_name) or f"bank_statement_{index}.xlsx"
		bank_input_path = run_dir / f"input_{index:02d}_{bank_safe_name}"
		bank_storage.save(bank_input_path)

		bank_workbook = load_workbook(bank_input_path)
		bank_sheet = pick_bank_statement_sheet(bank_workbook)
		header_row = 2

		for row_number in range(1, min(bank_sheet.max_row, 10) + 1):
			row_values = [normalize_text(bank_sheet.cell(row_number, column).value) for column in range(1, 6)]
			if "DATE" in row_values and "TRANSACTIONS" in row_values and "CREDIT" in row_values:
				header_row = row_number
				break

		for row_number in range(header_row + 1, bank_sheet.max_row + 1):
			transaction_cell = bank_sheet.cell(row=row_number, column=2)
			credit_cell = bank_sheet.cell(row=row_number, column=4)
			transaction_info = parse_speedpoint_bank_transaction(transaction_cell.value)
			credit_amount = parse_amount(credit_cell.value)
			if not transaction_info or credit_amount is None:
				continue

			expected_amount = float(credit_amount)
			match_row: int | None = None
			match_amount: float | None = None
			exact_match = False
			reason = "No match found"

			if transaction_info["transaction_type"] == "NEDLNK DP":
				match_row, match_amount, exact_match = find_speedpoint_total_match(
					speedpoint_sheet,
					expected_amount,
					start_column=7,
					end_column=15,
				)
				if exact_match and match_row is not None:
					highlight_speedpoint_range(speedpoint_sheet, match_row, 7, 15)
					highlight_bank_match(bank_sheet, row_number, 4)
					total_matches += 1
					continue
				if match_row is not None:
					reason = "Amount mismatch"
				else:
					reason = "No matching G:O total"
			else:
				terminal_column = int(transaction_info["speedpoint_column"])
				match_row, match_amount, exact_match = find_speedpoint_match(
					speedpoint_sheet,
					terminal_column,
					int(transaction_info["batch_number"]),
					expected_amount,
				)
				if exact_match and match_row is not None:
					highlight_speedpoint_match(speedpoint_sheet, match_row, terminal_column)
					highlight_bank_match(bank_sheet, row_number, 4)
					total_matches += 1
					continue
				reason = "Batch not found"
				if match_row is not None:
					reason = "Amount mismatch"
			unresolved_rows.append(
				{
					"file_name": bank_original_name,
					"sheet_name": bank_sheet.title,
					"row_number": row_number,
					"transaction": compact_text(transaction_cell.value),
					"credit": expected_amount,
					"terminal": transaction_info.get("speedpoint_terminal") or transaction_info.get("bank_terminal") or "",
					"bank_terminal": transaction_info["bank_terminal"],
					"batch_number": transaction_info["batch_number"],
					"transaction_type": transaction_info["transaction_type"],
					"reason": reason,
					"matched_row": match_row,
					"matched_amount": match_amount,
				}
			)

		bank_output_name = f"edited_{index:02d}_{bank_safe_name}"
		bank_output_path = run_dir / bank_output_name
		bank_workbook.save(bank_output_path)
		bank_download_name = format_download_name(bank_original_name)
		edited_files.append(
			{
				"download_name": bank_download_name,
				"stored_name": bank_output_name,
				"kind": "bank",
			}
		)

	speedpoint_output_name = f"edited_{speedpoint_safe_name}"
	speedpoint_output_path = run_dir / speedpoint_output_name
	speedpoint_workbook.save(speedpoint_output_path)
	edited_files.insert(
		0,
		{
			"download_name": format_download_name(speedpoint_original_name, strip_leading_number=True),
			"stored_name": speedpoint_output_name,
			"kind": "speedpoint",
		},
	)

	manifest = {
		"run_id": run_id,
		"speedpoint_original_name": speedpoint_original_name,
		"speedpoint_download_name": format_download_name(speedpoint_original_name, strip_leading_number=True),
		"bank_count": len(bank_storages),
		"matched_count": total_matches,
		"unresolved_count": len(unresolved_rows),
		"unresolved_rows": unresolved_rows,
		"edited_files": edited_files,
		"zip_filename": f"speedpoint_reconciliation_{run_id[:8]}.zip",
	}
	manifest_path = run_dir / "manifest.json"
	manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
	zip_path = build_result_zip(run_dir, manifest)
	manifest["zip_path"] = zip_path.name
	manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
	return manifest


def load_speedpoint_manifest(run_id: str) -> dict[str, object] | None:
	manifest_path = SPEEDPOINT_RUNS_DIR / run_id / "manifest.json"
	if not manifest_path.exists():
		return None
	return json.loads(manifest_path.read_text(encoding="utf-8"))


def stage_workbook(
    batch_id: int,
    file_storage,
    original_filename: str,
    file_date: str,
) -> tuple[int | None, int, bool]:
    source_path = Path(file_storage.filename)
    safe_name = secure_filename(source_path.name) or "upload.xlsx"
    temp_path = UPLOAD_DIR / f"batch{batch_id}_{safe_name}"
    file_storage.save(temp_path)

    try:
        file_hash = sha256_file(temp_path)
        existing = query_one("SELECT id FROM import_files WHERE file_hash = ?", (file_hash,))
        if existing:
            temp_path.unlink(missing_ok=True)
            return None, 0, True

        stored_name = f"{file_hash[:12]}_{safe_name}"
        stored_path = UPLOAD_DIR / stored_name
        temp_path.replace(stored_path)

        cursor = execute(
            """
            INSERT INTO import_files (batch_id, original_filename, stored_filename, file_hash, file_date)
            VALUES (?, ?, ?, ?, ?)
            """,
            (batch_id, original_filename, stored_name, file_hash, file_date),
        )
        file_id = cursor.lastrowid
        workbook = load_workbook(stored_path, data_only=True, read_only=True)
        worksheet = pick_payment_sheet(workbook)
        staged_rows = 0

        for row_number in PAYMENT_RANGE:
            supplier_name = worksheet.cell(row=row_number, column=1).value
            if supplier_name in (None, ""):
                continue
            supplier_name = str(supplier_name).strip()
            if not supplier_name:
                continue
            supplier_key = normalize_text(supplier_name)
            for column_number in PAYMENT_COLUMNS:
                amount = parse_money(worksheet.cell(row=row_number, column=column_number).value)
                if amount is None:
                    continue
                duplicate_row = query_one(
                    """
                    SELECT 1
                    FROM payments
                    WHERE payment_date = ?
                      AND supplier_name = ?
                      AND amount_paid = ?
                    UNION
                    SELECT 1
                    FROM staged_payments
                    WHERE payment_date = ?
                      AND supplier_name = ?
                      AND amount_paid = ?
                    LIMIT 1
                    """,
                    (
                        file_date,
                        supplier_name,
                        amount,
                        file_date,
                        supplier_name,
                        amount,
                    ),
                )
                if duplicate_row:
                    continue
                account = resolve_account_for_supplier(supplier_name)
                account_id = account["account_id"] if account and "account_id" in account.keys() else None
                if account is not None and "id" in account.keys():
                    account_id = account["id"]
                execute(
                    """
                    INSERT INTO staged_payments (
                        batch_id, file_id, payment_date, supplier_name, supplier_key,
                        amount_paid, source_sheet, source_row, source_column, account_id
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT (file_id, source_sheet, source_row, source_column) DO NOTHING
                    """,
                    (
                        batch_id,
                        file_id,
                        file_date,
                        supplier_name,
                        supplier_key,
                        amount,
                        worksheet.title,
                        row_number,
                        column_number,
                        account_id,
                    ),
                )
                staged_rows += 1

        finalize_resolved_rows(batch_id)
        return file_id, staged_rows, False
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def date_to_iso(value: object) -> str | None:
	if value is None:
		return None
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, date):
		return value.isoformat()
	if isinstance(value, str):
		for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y%m%d", "%Y%m%d%H%M%S"):
			try:
				return datetime.strptime(value, fmt).date().isoformat()
			except ValueError:
				continue
	return None


def parse_ofx_transactions(file_path: Path) -> list[dict[str, object]]:
	raw_text = file_path.read_text(encoding="latin-1")
	raw_text = re.sub(r"CHARSET:NONE", "CHARSET:1252", raw_text, flags=re.IGNORECASE)
	with StringIO(raw_text) as handle:
		ofx = OfxParser.parse(handle)

	transactions = []

	account_list = getattr(ofx, "accounts", None) or []
	if account_list:
		for account in account_list:
			statement = getattr(account, "statement", None)
			if statement:
				transactions.extend(getattr(statement, "transactions", []) or [])
	else:
		statement = getattr(ofx, "statement", None)
		if statement:
			transactions.extend(getattr(statement, "transactions", []) or [])

	parsed = []
	for index, txn in enumerate(transactions, start=1):
		description = ""
		for attr in ("payee", "memo", "name", "type"):
			value = getattr(txn, attr, None)
			if value:
				description = str(value).strip()
				break
		amount = parse_money(getattr(txn, "amount", None))
		payment_date = date_to_iso(getattr(txn, "date", None)) or date.today().isoformat()
		if not description or amount is None:
			continue
		if amount >= 0:
			continue
		parsed.append(
			{
				"index": index,
				"description": description,
				"amount": abs(amount),
				"payment_date": payment_date,
			}
		)
	return parsed


def stage_ofx_file(
	batch_id: int,
	file_storage,
	original_filename: str,
) -> tuple[int | None, int, bool]:
	source_path = Path(file_storage.filename)
	safe_name = secure_filename(source_path.name) or "statement.ofx"
	temp_path = UPLOAD_DIR / f"batch{batch_id}_{safe_name}"
	file_storage.save(temp_path)

	try:
		file_hash = sha256_file(temp_path)
		existing = query_one("SELECT id FROM import_files WHERE file_hash = ?", (file_hash,))
		if existing:
			temp_path.unlink(missing_ok=True)
			return None, 0, True

		transactions = parse_ofx_transactions(temp_path)
		file_date = transactions[0]["payment_date"] if transactions else date.today().isoformat()

		stored_name = f"{file_hash[:12]}_{safe_name}"
		stored_path = UPLOAD_DIR / stored_name
		temp_path.replace(stored_path)

		cursor = execute(
			"""
			INSERT INTO import_files (batch_id, original_filename, stored_filename, file_hash, file_date)
			VALUES (?, ?, ?, ?, ?)
			""",
			(batch_id, original_filename, stored_name, file_hash, file_date),
		)
		file_id = cursor.lastrowid
		staged_rows = 0

		for txn in transactions:
			description = str(txn["description"]).strip()
			amount = parse_money(txn["amount"])
			payment_date = str(txn["payment_date"])
			if not description or amount is None:
				continue
			account = resolve_account_for_supplier(description)
			account_id = account["id"] if account is not None and "id" in account.keys() else None
			if account is not None and "account_id" in account.keys():
				account_id = account["account_id"]
			execute(
				"""
				INSERT INTO staged_payments (
					batch_id, file_id, payment_date, supplier_name, supplier_key,
					amount_paid, source_sheet, source_row, source_column, account_id
				)
				VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
				ON CONFLICT (file_id, source_sheet, source_row, source_column) DO NOTHING
				""",
				(
					batch_id,
					file_id,
					payment_date,
					description,
					normalize_text(description),
					amount,
					"OFX",
					int(txn["index"]),
					1,
					account_id,
				),
			)
			staged_rows += 1

		finalize_resolved_rows(batch_id)
		return file_id, staged_rows, False
	except Exception:
		temp_path.unlink(missing_ok=True)
		raise


def finalize_resolved_rows(batch_id: int) -> int:
	cursor = execute(
		"""
		INSERT INTO payments (
			payment_date, supplier_name, amount_paid, account_id,
			source_file_name, source_file_hash, source_sheet, source_row, source_column
		)
		SELECT
			staged_payments.payment_date,
			staged_payments.supplier_name,
			staged_payments.amount_paid,
			staged_payments.account_id,
			import_files.original_filename,
			import_files.file_hash,
			staged_payments.source_sheet,
			staged_payments.source_row,
			staged_payments.source_column
		FROM staged_payments
		JOIN import_files ON import_files.id = staged_payments.file_id
		WHERE staged_payments.batch_id = ?
		  AND staged_payments.account_id IS NOT NULL
		ON CONFLICT (source_file_hash, source_sheet, source_row, source_column) DO NOTHING
		""",
		(batch_id,),
	)
	return cursor.rowcount if cursor.rowcount != -1 else 0


def latest_batch_id(source_type: str | None = None) -> int | None:
	if source_type:
		row = query_one(
			"SELECT id FROM import_batches WHERE source_type = ? ORDER BY id DESC LIMIT 1",
			(source_type,),
		)
	else:
		row = query_one("SELECT id FROM import_batches ORDER BY id DESC LIMIT 1")
	return row["id"] if row else None


def normalize_month(value: str | None) -> str:
	if not value:
		return ""
	try:
		return datetime.strptime(value, "%Y-%m").strftime("%Y-%m")
	except ValueError:
		return ""


def latest_payment_month() -> str | None:
	row = query_one(
		"""
		SELECT substr(payment_date, 1, 7) AS month_value
		FROM payments
		WHERE payment_date IS NOT NULL
		ORDER BY payment_date DESC
		LIMIT 1
		"""
	)
	return row["month_value"] if row and row["month_value"] else None


def available_payment_months() -> list[sqlite3.Row]:
	return query_all(
		"""
		SELECT substr(payment_date, 1, 7) AS month_value
		FROM payments
		WHERE payment_date IS NOT NULL
		GROUP BY month_value
		ORDER BY month_value DESC
		"""
	)


def available_payment_sources() -> list[str]:
	rows = query_all(
		"""
		SELECT DISTINCT source_type
		FROM import_batches
		WHERE source_type IS NOT NULL AND TRIM(source_type) <> ''
		ORDER BY source_type
		"""
	)
	return [row["source_type"] for row in rows]


def source_label(source_type: str | None) -> str:
	if source_type == "cash":
		return "Cash Payments"
	if source_type == "eft":
		return "EFT Payments"
	if not source_type:
		return "All sources"
	return source_type.replace("_", " ").title()


def available_account_source_options() -> list[tuple[str, str]]:
	values = query_all(
		"""
		SELECT DISTINCT source_sheet AS source_name
		FROM accounts
		WHERE source_sheet IS NOT NULL AND TRIM(source_sheet) <> ''
		ORDER BY source_name
		"""
	)
	options = [("Manual", "Manual")]
	for row in values:
		source_name = row["source_name"]
		if source_name and source_name != "Manual":
			options.append((source_name, source_name))
	if len(options) == 1:
		options.extend([
			("Cash Payments", "Cash Payments"),
			("EFT Payments", "EFT Payments"),
		])
	return options


def month_label(month_value: str) -> str:
	return datetime.strptime(month_value, "%Y-%m").strftime("%B %Y")


def batch_summary(batch_id: int) -> dict:
	counts = query_one(
		"""
		SELECT
			COUNT(*) AS staged_total,
			SUM(CASE WHEN account_id IS NULL THEN 1 ELSE 0 END) AS unresolved_total,
			SUM(CASE WHEN account_id IS NOT NULL THEN 1 ELSE 0 END) AS resolved_total
		FROM staged_payments
		WHERE batch_id = ?
		""",
		(batch_id,),
	)
	files = query_all(
		"""
		SELECT original_filename, file_date, stored_filename
		FROM import_files
		WHERE batch_id = ?
		ORDER BY id DESC
		""",
		(batch_id,),
	)
	unresolved = query_all(
		"""
		SELECT
			staged_payments.id AS staged_id,
			staged_payments.batch_id,
			staged_payments.payment_date,
			staged_payments.supplier_name,
			staged_payments.supplier_key,
			staged_payments.amount_paid,
			staged_payments.source_sheet,
			staged_payments.source_row,
			staged_payments.source_column,
			import_files.original_filename,
			import_files.file_date
		FROM staged_payments
		JOIN import_files ON import_files.id = staged_payments.file_id
		WHERE staged_payments.account_id IS NULL
		ORDER BY import_files.file_date DESC, import_files.original_filename DESC,
				 staged_payments.source_row ASC, staged_payments.source_column ASC
		""",
	)
	return {
		"counts": counts,
		"files": files,
		"unresolved": unresolved,
	}


def account_summary_for_month(month_value: str, source_type: str | None = None) -> list[sqlite3.Row]:
	params: list[object] = [month_value]
	source_clause = ""
	if source_type and source_type != "all":
		source_clause = " AND import_batches.source_type = ?"
		params.append(source_type)
	return query_all(
		f"""
		SELECT
			accounts.id AS account_id,
			accounts.name AS account_name,
			SUM(payments.amount_paid) AS amount_paid
		FROM payments
		JOIN accounts ON accounts.id = payments.account_id
		JOIN import_files ON import_files.file_hash = payments.source_file_hash
		JOIN import_batches ON import_batches.id = import_files.batch_id
		WHERE substr(payments.payment_date, 1, 7) = ?{source_clause}
		GROUP BY accounts.id, accounts.name
		ORDER BY accounts.name
		""",
		tuple(params),
	)


def account_transactions_for_month(month_value: str, account_id: int, source_type: str | None = None) -> list[sqlite3.Row]:
	params: list[object] = [month_value, account_id]
	source_clause = ""
	if source_type and source_type != "all":
		source_clause = " AND import_batches.source_type = ?"
		params.append(source_type)
	return query_all(
		f"""
		SELECT
			payments.payment_date,
			payments.supplier_name,
			payments.amount_paid,
			accounts.name AS account_name,
			payments.source_file_name,
			payments.source_sheet,
			payments.source_row,
			payments.source_column,
			import_batches.source_type AS source_type
		FROM payments
		JOIN accounts ON accounts.id = payments.account_id
		JOIN import_files ON import_files.file_hash = payments.source_file_hash
		JOIN import_batches ON import_batches.id = import_files.batch_id
		WHERE substr(payments.payment_date, 1, 7) = ?
		  AND payments.account_id = ?{source_clause}
		ORDER BY payments.payment_date, payments.id
		""",
		tuple(params),
	)


def bootstrap_storage() -> None:
	with app.app_context():
		initialize_storage()
		sync_account_names()
		app._seeded_accounts = True


@app.before_request
def prepare_app() -> None:
	initialize_storage()
	if not getattr(app, "_seeded_accounts", False):
		sync_account_names()
		app._seeded_accounts = True


@app.teardown_appcontext
def close_connection(_: Exception | None = None) -> None:
	db = g.pop("db", None)
	if db is not None:
		db.close()


@app.route("/")
def index():
	return render_template(
		"index.html",
		show_nav=False,
	)


@app.route("/expenses")
def expenses():
	return render_template("expenses.html", title="Expenses")


@app.route("/reports")
def reports():
	month_value = normalize_month(request.args.get("month")) or latest_payment_month()
	return render_template(
		"reports.html",
		title="Reports",
		month_value=month_value,
	)


@app.route("/speed-point-control-recon/uploads", methods=["GET", "POST"])
def speedpoint_uploads():
	if request.method == "POST":
		speedpoint_file = request.files.get("speedpoint_file")
		bank_files = [file_storage for file_storage in request.files.getlist("bank_files") if file_storage and file_storage.filename]

		if not speedpoint_file or not speedpoint_file.filename:
			flash("Please choose one Speedpoints .xlsx file.", "warning")
			return redirect(url_for("speedpoint_uploads"))

		if Path(speedpoint_file.filename).suffix.lower() not in SPEEDPOINT_UPLOAD_EXTENSIONS:
			flash("The Speedpoints file must be a .xlsx workbook.", "warning")
			return redirect(url_for("speedpoint_uploads"))

		if not bank_files:
			flash("Please choose at least one Bank Statements .xlsx file.", "warning")
			return redirect(url_for("speedpoint_uploads"))

		for bank_file in bank_files:
			if Path(bank_file.filename).suffix.lower() not in SPEEDPOINT_UPLOAD_EXTENSIONS:
				flash("All Bank Statements files must be .xlsx workbooks.", "warning")
				return redirect(url_for("speedpoint_uploads"))

		manifest = process_speedpoint_reconciliation(speedpoint_file, bank_files)
		return redirect(url_for("speedpoint_uploads", run_id=manifest["run_id"]))

	run_id = request.args.get("run_id", type=str)
	results = load_speedpoint_manifest(run_id) if run_id else None
	if run_id and not results:
		flash("That reconciliation result is no longer available.", "warning")
		return redirect(url_for("speedpoint_uploads"))

	return render_template(
		"speedpoint_uploads.html",
		title="Uploads",
		active_page="uploads",
		show_nav=False,
		run_id=run_id,
		results=results,
		download_url=url_for("speedpoint_download", run_id=run_id) if run_id else None,
	)


@app.route("/speed-point-control-recon/uploads/download/<run_id>")
def speedpoint_download(run_id: str):
	manifest = load_speedpoint_manifest(run_id)
	if not manifest:
		flash("The edited files are no longer available.", "warning")
		return redirect(url_for("speedpoint_uploads"))

	zip_path = SPEEDPOINT_RUNS_DIR / run_id / manifest["zip_filename"]
	if not zip_path.exists():
		flash("The download bundle could not be found.", "warning")
		return redirect(url_for("speedpoint_uploads", run_id=run_id))

	return send_file(
		zip_path,
		as_attachment=True,
		download_name=manifest["zip_filename"],
	)


@app.route("/cash-payments", methods=["GET"])
def cash_payments():
	batch_id = request.args.get("batch_id", type=int) or latest_batch_id("cash")
	summary = batch_summary(batch_id) if batch_id else None
	accounts = query_all("SELECT id, name, source_sheet FROM accounts ORDER BY name")
	return render_template(
		"cash_payments.html",
		batch_id=batch_id,
		summary=summary,
		accounts=accounts,
		available_account_source_options=available_account_source_options(),
	)


@app.route("/cash-payments/upload", methods=["POST"])
def upload_cash_payments():
    files = request.files.getlist("files")
    if not files:
        flash("Please choose one or more Excel files.", "warning")
        return redirect(url_for("cash_payments"))

    batch_id = execute(
        "INSERT INTO import_batches (source_folder, source_type, status) VALUES (?, ?, ?)",
        (request.form.get("source_folder", ""), "cash", "open"),
    ).lastrowid

    imported_files = 0
    staged_rows = 0
    duplicates = 0

    for file_storage in files:
        if not file_storage or not file_storage.filename:
            continue
        if not allowed_file(file_storage.filename):
            flash(f"Skipped {file_storage.filename}: only .xlsx and .xlsm files are supported.", "warning")
            continue
        try:
            file_date = parse_filename_date(Path(file_storage.filename).name)
        except ValueError as exc:
            flash(str(exc), "danger")
            continue

        file_id, staged_count, duplicate_flag = stage_workbook(
            batch_id=batch_id,
            file_storage=file_storage,
            original_filename=Path(file_storage.filename).name,
            file_date=file_date,
        )
        if duplicate_flag:
            duplicates += 1
            continue
        if file_id is not None:
            imported_files += 1
            staged_rows += staged_count

    if imported_files == 0 and staged_rows == 0:
        if duplicates:
            flash(f"All {duplicates} uploaded file(s) were already imported.", "info")
        else:
            flash("No new workbook rows were imported.", "info")
    else:
        message = f"Imported {imported_files} new file(s) and staged {staged_rows} payment row(s)."
        if duplicates:
            message += f" {duplicates} file(s) were already in the database."
        flash(message, "success")

    return redirect(url_for("cash_payments", batch_id=batch_id))


@app.route("/eft-payments", methods=["GET"])
def eft_payments():
	batch_id = request.args.get("batch_id", type=int) or latest_batch_id("eft")
	summary = batch_summary(batch_id) if batch_id else None
	accounts = query_all("SELECT id, name, source_sheet FROM accounts ORDER BY name")
	return render_template(
		"eft_payments.html",
		batch_id=batch_id,
		summary=summary,
		accounts=accounts,
		available_account_source_options=available_account_source_options(),
	)


@app.route("/eft-payments/upload", methods=["POST"])
def upload_eft_payments():
	files = request.files.getlist("files")
	if not files:
		flash("Please choose one or more OFX files.", "warning")
		return redirect(url_for("eft_payments"))

	batch_id = execute(
		"INSERT INTO import_batches (source_folder, source_type, status) VALUES (?, ?, ?)",
		(request.form.get("source_folder", ""), "eft", "open"),
	).lastrowid

	imported_files = 0
	staged_rows = 0
	duplicates = 0

	for file_storage in files:
		if not file_storage or not file_storage.filename:
			continue
		if Path(file_storage.filename).suffix.lower() != ".ofx":
			flash(f"Skipped {file_storage.filename}: only .ofx files are supported.", "warning")
			continue
		file_id, staged_count, duplicate_flag = stage_ofx_file(
			batch_id=batch_id,
			file_storage=file_storage,
			original_filename=Path(file_storage.filename).name,
		)
		if duplicate_flag:
			duplicates += 1
			continue
		if file_id is not None:
			imported_files += 1
			staged_rows += staged_count

	if imported_files == 0 and staged_rows == 0:
		if duplicates:
			flash(f"All {duplicates} uploaded file(s) were already imported.", "info")
		else:
			flash("No new OFX transactions were imported.", "info")
	else:
		message = f"Imported {imported_files} new file(s) and staged {staged_rows} transaction(s)."
		if duplicates:
			message += f" {duplicates} file(s) were already in the database."
		flash(message, "success")

	return redirect(url_for("eft_payments", batch_id=batch_id))


def _resolve_payment_item(staged_id: int, redirect_endpoint: str):
	staged = query_one(
		"""
		SELECT staged_payments.*, import_files.original_filename, import_files.file_date
		FROM staged_payments
		JOIN import_files ON import_files.id = staged_payments.file_id
		WHERE staged_payments.id = ?
		""",
		(staged_id,),
	)
	if not staged:
		flash("That transaction could not be found.", "warning")
		return redirect(url_for("cash_payments"))

	if staged["account_id"] is not None:
		flash("That transaction is already resolved.", "info")
		return redirect(url_for(redirect_endpoint))

	selected_account_id = request.form.get("account_id", type=int)
	new_account_name = request.form.get("new_account_name", "").strip()
	new_account_source = request.form.get("new_account_source", "Manual").strip()
	account_id = None
	account_name = None

	if selected_account_id:
		account = query_one("SELECT id FROM accounts WHERE id = ?", (selected_account_id,))
		if account:
			account_id = selected_account_id
			account_name = query_one("SELECT name FROM accounts WHERE id = ?", (selected_account_id,))["name"]
	elif new_account_name:
		existing = find_account_by_normalized_name(new_account_name)
		if existing:
			account_id = existing["id"]
			account_name = existing["name"]
		else:
			upper_name = canonical_account_name(new_account_name)
			source_sheet = None if new_account_source == "Manual" else new_account_source
			cursor = execute("INSERT INTO accounts (name, source_sheet) VALUES (?, ?) ON CONFLICT (name) DO NOTHING", (upper_name, source_sheet))
			account_id = cursor.lastrowid
			account_name = upper_name

	if not account_id:
		flash("Choose an existing account or type a new account name before saving.", "warning")
		return redirect(url_for(redirect_endpoint))

	ensure_rule(staged["supplier_name"], account_id)
	execute(
		"""
		UPDATE staged_payments
		SET account_id = ?
		WHERE batch_id = ? AND supplier_key = ? AND account_id IS NULL
		""",
		(account_id, staged["batch_id"], staged["supplier_key"]),
	)
	finalize_resolved_rows(staged["batch_id"])
	flash(f"Saved '{staged['supplier_name']}' to '{account_name}'.", "success")
	return redirect(url_for(redirect_endpoint))


@app.route("/cash-payments/resolve/<int:staged_id>", methods=["POST"])
def resolve_cash_payment_item(staged_id: int):
	return _resolve_payment_item(staged_id, "cash_payments")


@app.route("/eft-payments/resolve/<int:staged_id>", methods=["POST"])
def resolve_eft_payment_item(staged_id: int):
	return _resolve_payment_item(staged_id, "eft_payments")


@app.route("/accounts", methods=["GET", "POST"])
def accounts():
    month_value = normalize_month(request.args.get("month")) or latest_payment_month()
    source_type = request.args.get("source") or "all"
    month_options = available_payment_months()
    source_options = [("all", "All sources")] + [
        (source_type_value, source_label(source_type_value)) for source_type_value in available_payment_sources()
    ]
    valid_source_types = {value for value, _ in source_options}
    if source_type not in valid_source_types:
        source_type = "all"
    if not month_value and month_options:
        month_value = month_options[0]["month_value"]
    rows = account_summary_for_month(month_value, source_type) if month_value else []
    return render_template(
        "accounts.html",
        month_value=month_value,
        month_label=month_label(month_value) if month_value else None,
        month_options=month_options,
        source_type=source_type,
        source_label=source_label(source_type),
        source_options=source_options,
        rows=rows,
    )


@app.route("/accounts/<int:account_id>", methods=["GET"])
def account_transactions(account_id: int):
    account = query_one("SELECT id, name FROM accounts WHERE id = ?", (account_id,))
    if not account:
        flash("That account could not be found.", "warning")
        return redirect(url_for("accounts"))

    month_value = normalize_month(request.args.get("month")) or latest_payment_month()
    source_type = request.args.get("source") or "all"
    month_options = available_payment_months()
    source_options = [("all", "All sources")] + [
        (source_type_value, source_label(source_type_value)) for source_type_value in available_payment_sources()
    ]
    valid_source_types = {value for value, _ in source_options}
    if source_type not in valid_source_types:
        source_type = "all"
    if not month_value and month_options:
        month_value = month_options[0]["month_value"]
    transactions = account_transactions_for_month(month_value, account_id, source_type) if month_value else []
    total_amount = sum(row["amount_paid"] for row in transactions)
    return render_template(
        "account_transactions.html",
        account_name=account["name"],
        account_id=account_id,
        month_value=month_value,
        month_label=month_label(month_value) if month_value else None,
        month_options=month_options,
        source_type=source_type,
        source_label=source_label(source_type),
        source_options=source_options,
        transactions=transactions,
        total_amount=total_amount,
    )


@app.route("/settings", methods=["GET", "POST"])
def settings():
	if request.method == "POST":
		action = request.form.get("action", "create_account").strip()
		if action == "update_account":
			account_id = request.form.get("account_id", type=int)
			account_name = request.form.get("account_name", "").strip()
			source_value = request.form.get("account_source", "Manual").strip()
			if not account_id:
				flash("Could not update the selected account.", "danger")
			else:
				existing = query_one("SELECT id FROM accounts WHERE id = ?", (account_id,))
				if not existing:
					flash("That account no longer exists.", "warning")
				elif not account_name:
					flash("Account name cannot be empty.", "warning")
				else:
					upper_name = canonical_account_name(account_name)
					duplicate = find_account_by_normalized_name(upper_name, exclude_account_id=account_id)
					if duplicate:
						flash(f"Account '{upper_name}' already exists.", "info")
					else:
						source_sheet = None if source_value == "Manual" else source_value
						execute(
							"UPDATE accounts SET name = ?, source_sheet = ? WHERE id = ?",
							(upper_name, source_sheet, account_id),
						)
						flash(f"Updated account '{upper_name}'.", "success")
		elif action == "delete_account":
			account_id = request.form.get("account_id", type=int)
			if not account_id:
				flash("Could not delete the selected account.", "danger")
			else:
				account = query_one("SELECT name FROM accounts WHERE id = ?", (account_id,))
				if not account:
					flash("That account no longer exists.", "warning")
				else:
					try:
						execute("DELETE FROM accounts WHERE id = ?", (account_id,))
					except DB_INTEGRITY_ERROR:
						flash(
							f"Cannot delete account '{account['name']}' because it is still used by payments.",
							"warning",
						)
					else:
						flash(f"Deleted account '{account['name']}'.", "success")
		else:
			account_name = request.form.get("account_name", "").strip()
			source_value = request.form.get("account_source", "Manual").strip()
			if account_name:
				existing = find_account_by_normalized_name(account_name)
				if existing:
					flash(f"Account '{account_name}' already exists.", "info")
				else:
					upper_name = canonical_account_name(account_name)
					source_sheet = None if source_value == "Manual" else source_value
					execute(
						"INSERT INTO accounts (name, source_sheet) VALUES (?, ?) ON CONFLICT (name) DO NOTHING",
						(upper_name, source_sheet),
					)
					flash(f"Created account '{upper_name}'.", "success")
		return redirect(url_for("settings"))

	account_rows = query_all(
		"""
		SELECT accounts.id, accounts.name, accounts.source_sheet,
		       COUNT(DISTINCT supplier_rules.id) AS linked_suppliers
		FROM accounts
		LEFT JOIN supplier_rules ON supplier_rules.account_id = accounts.id
		GROUP BY accounts.id
		ORDER BY accounts.name
		"""
	)
	rules = query_all(
		"""
		SELECT supplier_rules.supplier_name, accounts.name AS account_name
		FROM supplier_rules
		JOIN accounts ON accounts.id = supplier_rules.account_id
		ORDER BY supplier_rules.supplier_name
		"""
	)
	return render_template(
		"settings.html",
		accounts=account_rows,
		rules=rules,
		available_account_source_options=available_account_source_options(),
	)


if __name__ == "__main__":
	with app.app_context():
		initialize_storage()
		sync_account_names()
	app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")), debug=os.environ.get("FLASK_DEBUG", "0") == "1")
